import pytest
from openpyxl import load_workbook

def get_test_data_from_excel(file_path, sheet_name):
    """
    从 Excel 文件中读取测试数据
    """
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    test_data = []
    test_data2 = ""
    for row in sheet.iter_rows(min_row=2, values_only=True):
        test_data.append(row)
        test_data.append(list(test_data[0]))
        test_data.pop(0)
        for i in test_data:
            test_data = i
    print(test_data)
    print(test_data[2])

    return test_data

if __name__ == '__main__':
    get_test_data_from_excel('../data/site_test.xlsx','Sheet1')
# @pytest.mark.parametrize("username, password", get_test_data_from_excel("test_data.xlsx", "login"))
# def test_login(username, password):
#     """
#     登录测试
#     """
#     # 执行登录操作并进行断言
#     assert login(username, password)